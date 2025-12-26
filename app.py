import logging
from flask import Flask, render_template, request, redirect, url_for
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# ========================
# Configuración de logging
# ========================
logging.basicConfig(
    filename='log.txt',          # Archivo donde se guardan los logs
    level=logging.INFO,          # Nivel de información mínimo
    format='%(asctime)s - %(levelname)s - %(message)s'
)

logging.info("La app Flask se ha iniciado")

# ========================
# Inicializar Flask
# ========================
app = Flask(__name__)

ARCHIVO = "registros_kayak.xlsx"

# ---------------- CREAR EXCEL ----------------
def crear_excel():
    if not os.path.exists(ARCHIVO):
        wb = Workbook()
        ws = wb.active
        ws.append(["Fecha", "Hora", "Tipo", "Valor"])
        wb.save(ARCHIVO)
        logging.info("Archivo Excel creado")

# ---------------- LEER REGISTROS ----------------
def leer_registros():
    wb = load_workbook(ARCHIVO)
    ws = wb.active
    registros = []

    for fila in ws.iter_rows(min_row=2, values_only=True):
        registros.append({
            "fecha": fila[0],
            "hora": fila[1],
            "tipo": fila[2],
            "valor": fila[3]
        })

    logging.info(f"{len(registros)} registros leídos del Excel")
    return registros

# ---------------- GUARDAR TODO ----------------
def guardar_todo(registros):
    wb = Workbook()
    ws = wb.active
    ws.append(["Fecha", "Hora", "Tipo", "Valor"])

    for r in registros:
        ws.append([r["fecha"], r["hora"], r["tipo"], r["valor"]])

    wb.save(ARCHIVO)
    logging.info(f"{len(registros)} registros guardados en Excel")

# ---------------- INICIO ----------------
@app.route("/", methods=["GET", "POST"])
def index():
    crear_excel()
    registros = leer_registros()
    fecha_buscar = request.form.get("fecha_buscar", "")

    if fecha_buscar:
        registros = [r for r in registros if r["fecha"] == fecha_buscar]
        logging.info(f"Filtrado de registros por fecha: {fecha_buscar}, {len(registros)} encontrados")

    total = sum(r["valor"] for r in registros if isinstance(r["valor"], int))
    logging.info(f"Total calculado: {total}")

    return render_template(
        "index.html",
        registros=registros,
        total=total,
        fecha_buscar=fecha_buscar
    )

# ---------------- AGREGAR ----------------
@app.route("/agregar", methods=["POST"])
def agregar():
    tipo = request.form.get("tipo")
    valor = request.form.get("valor")
    valor = int(valor) if valor and valor.isdigit() else 0

    ahora = datetime.now()
    fecha = ahora.strftime("%Y-%m-%d")
    hora = ahora.strftime("%H:%M:%S")

    wb = load_workbook(ARCHIVO)
    ws = wb.active
    ws.append([fecha, hora, tipo, valor])
    wb.save(ARCHIVO)

    logging.info(f"Registro agregado: {fecha} {hora} {tipo} {valor}")
    return redirect(url_for("index"))

# ---------------- ELIMINAR ----------------
@app.route("/eliminar/<int:index>")
def eliminar(index):
    registros = leer_registros()
    if 0 <= index < len(registros):
        eliminado = registros.pop(index)
        guardar_todo(registros)
        logging.info(f"Registro eliminado: {eliminado}")
    return redirect(url_for("index"))

# ---------------- EDITAR ----------------
@app.route("/editar/<int:index>", methods=["GET", "POST"])
def editar(index):
    registros = leer_registros()
    registro = registros[index]

    if request.method == "POST":
        registro["fecha"] = request.form["fecha"]
        registro["hora"] = request.form["hora"]
        registro["tipo"] = request.form["tipo"]

        valor = request.form["valor"]
        registro["valor"] = int(valor) if valor.isdigit() else 0

        guardar_todo(registros)
        logging.info(f"Registro editado: {registro}")
        return redirect(url_for("index"))

    return render_template("editar.html", r=registro, index=index)

# ---------------- RUTA DE LOGS ----------------
@app.route("/logs")
def ver_logs():
    try:
        with open("log.txt", "r", encoding="utf-8") as f:
            contenido = f.read()
        return f"<pre>{contenido}</pre>"
    except Exception as e:
        return f"No se pudo leer el archivo de logs: {e}"

# ---------------- MAIN ----------------
if __name__ == "__main__":
    logging.info("Servidor iniciado en 0.0.0.0:5000")
    app.run(debug=True, host="0.0.0.0", port=5000)
